"""
Модуль создания результатов
"""
import json
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from database.db_operations import (
    AnswerTest,
    AnswerTestTotalResult
)

# ДА! Захардкодил. Лень было новую таблицу делать
MOTIVES = {
    '1': 'Мотив вознаграждения – вы работаете ради денег и других благ',
    '2': 'Социальный мотив – вам важно одобрение руководства и коллектива',
    '3': 'Процессный мотив – вы трудитесь ради удовольствия от самого процесса работы',
    '4': 'Мотив достижения – вы стремитесь к самоутверждению и самореализации',
    '5': 'Идейный мотив – вам важно достижение совместных с компанией высоких целей'
}


class UserResult:
    def __init__(self, tg_id):
        self.tg_id = tg_id

    def _save_db_result(self):
        self.db_result = AnswerTest().user_result(self.tg_id)

    def main_motivation_find(self):
        self._save_db_result()
        sum_answer = {'1': 0,
                      '2': 0,
                      '3': 0,
                      '4': 0,
                      '5': 0}
        for record in self.db_result:
            if record['number'] >= 6:
                answer_dict = json.loads(record['answer'])
                for key, value in answer_dict.items():
                    sum_answer[key] += int(value)

        min_key = min(sum_answer, key=lambda x: sum_answer[x])

        return MOTIVES[min_key]

class TotalResul:

    def total_user(self) -> dict:
        """
        Общее количество пользователей

        Возвращает словарь {'total_user': ...}
        """
        result = AnswerTestTotalResult().total_count_user()
        return result[0]['total_user']

    def group_category(self):
        """
            Возвращает ответы 1-5 по категориям
        """
        result_dict = {
            1: {},
            2: {},
            3: {},
            4: {},
            5: {}
        }
        for key in result_dict:
            result_answer = AnswerTestTotalResult().group_category(key)
            result_dict[key] = {
                record['answer']: record['total_user']
                for record in result_answer
            }
        return result_dict

    def total_count_priority(self):
        """Получить ответы пользователей"""
        return AnswerTestTotalResult().total_count_priority()


class ExelResult:
    @property
    def file_name(self):
        return 'service/result.xlsx'

    def fill_first_sheet(self):
        """
        Первый лист - результаты
        """
        self.ready_users = list(AnswerTestTotalResult().ready_users().values())

        self.sheet1 = self.workbook.create_sheet(title="Результаты")

        offset = 0
        while True:
            answers = AnswerTestTotalResult().format_result(
                pagination=(1500, offset)
            )
            print(answers)
            if not answers:
                break


    # region second
    def fill_category(self, number, question_text):
        """
        Заполнит результат первых 5 вопросов
        """
        first_row = number * 3
        cell = self.sheet2.cell(
            row=first_row,
            column=1,
            value=question_text
        )
        cell.font = Font(bold=True)
        # Получили весь словарь категорий
        sex_categories = list(self.group_category[number].keys())
        for index, category in enumerate(sex_categories, start=1):
            self.sheet2.cell(
                row=first_row + 1,
                column=index,
                value=category
            )
            self.sheet2.cell(
                row=first_row + 2,
                column=index,
                value=self.group_category[number].get(category, 0)
            )

    def fill_second_sheet(self):
        """
        Второй лист - агрегация статистики
        """
        self.sheet2 = self.workbook.create_sheet(title="Агрегация")

        tl = TotalResul()

        self.group_category = tl.group_category()
        print(self.group_category)

        # Всего пользователей
        cell = self.sheet2.cell(row=1, column=1, value='Всего пользователей:')
        cell.font = Font(bold=True)
        self.sheet2.cell(row=1, column=2, value=tl.total_user())

        # 1 Пол
        self.fill_category(1, 'Пол:')

        # 2 Возрастная категория
        self.fill_category(2, 'Возрастная категория')

        # 3 Муниципальный или городской округ проживания
        self.fill_category(3, 'Место проживания')

        # 4 Уровень образования
        self.fill_category(4, 'Уровень образования')

        # 5 Работа
        self.fill_category(5, 'Наличие работы')




        # В самом конце установим ширину колонок
        fixed_width = 30
        for i in range(1, 30):
            self.sheet2.column_dimensions[get_column_letter(i)].width = fixed_width

    # endregion



    def create_file_result(self):

        # Удаляем старый файл, если он существует
        if os.path.exists(self.file_name):
            os.remove(self.file_name)
            print(f"Старый файл '{self.file_name}' был удален.")

        # Создаем новую книгу
        self.workbook = openpyxl.Workbook()

        # Заполняем первый лист
        self.fill_first_sheet()

        # Заполняем второй лист
        self.fill_second_sheet()

        # Сохраняем файл
        self.workbook.save(self.file_name)




